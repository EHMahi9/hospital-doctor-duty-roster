from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy.orm import Session

from app.models.duty import DutyAssignment
from app.services.analytics_service import month_bounds


def roster_rows(db: Session, month: int, year: int) -> list[list[str]]:
    start, end = month_bounds(month, year)
    assignments = (
        db.query(DutyAssignment)
        .filter(DutyAssignment.duty_date >= start, DutyAssignment.duty_date <= end)
        .order_by(DutyAssignment.duty_date.asc(), DutyAssignment.duty_type.asc())
        .all()
    )
    return [
        [
            item.duty_date.isoformat(),
            item.duty_type.value,
            item.shift.value.title(),
            item.doctor.name,
            item.doctor.department.name if item.doctor.department else "",
            item.doctor.designation,
            "Manual" if item.is_manual_override else "Auto",
        ]
        for item in assignments
    ]


def export_roster_xlsx(db: Session, month: int, year: int) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = f"Roster {year}-{month:02d}"
    ws.append(["Date", "Duty Type", "Shift", "Doctor", "Department", "Designation", "Source"])
    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row in roster_rows(db, month, year):
        ws.append(row)

    widths = [14, 24, 14, 26, 22, 24, 12]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + index)].width = width

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def export_roster_pdf(db: Session, month: int, year: int) -> BytesIO:
    stream = BytesIO()
    document = SimpleDocTemplate(stream, pagesize=landscape(A4), rightMargin=24, leftMargin=24, topMargin=24, bottomMargin=24)
    styles = getSampleStyleSheet()
    story = [
        Paragraph(f"Hospital Doctor Duty Roster - {date(year, month, 1):%B %Y}", styles["Title"]),
        Spacer(1, 12),
    ]
    table_data = [["Date", "Duty Type", "Shift", "Doctor", "Department", "Designation", "Source"], *roster_rows(db, month, year)]
    table = Table(table_data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F172A")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    story.append(table)
    document.build(story)
    stream.seek(0)
    return stream
